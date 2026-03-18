#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Universal marketplace price parser for user-provided XLSX tables.

Usage:
  python marketplace/universal_price_parser.py --input input.xlsx --output output.xlsx
"""

from __future__ import annotations

import argparse
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import quote

UA = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
}

MARKET_DOMAIN = {
    "alibaba": "alibaba.com",
    "aliexpress": "aliexpress.com",
    "ebay": "ebay.com",
    "ozon": "ozon.ru",
    "wildberries": "wildberries.ru",
}

RUB = re.compile(r"(?:₽|руб|rub)", re.I)
USD = re.compile(r"(?:\$|usd)", re.I)
EUR = re.compile(r"(?:€|eur)", re.I)
PRICE_RX = [
    re.compile(r"(\d{1,3}(?:[\s\u00a0]\d{3})+(?:[\.,]\d{1,2})?)"),
    re.compile(r"(\d{2,8}(?:[\.,]\d{1,2})?)"),
]


def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def detect_currency(text: str) -> str:
    if RUB.search(text):
        return "RUB"
    if USD.search(text):
        return "USD"
    if EUR.search(text):
        return "EUR"
    return ""


def extract_prices(text: str) -> list[float]:
    vals = []
    for rx in PRICE_RX:
        for m in rx.finditer(text):
            raw = m.group(1).replace("\u00a0", " ").replace(" ", "").replace(",", ".")
            try:
                v = float(raw)
                if 1 < v < 100000000:
                    vals.append(v)
            except Exception:
                pass
    return vals


def likely_in_stock(text: str) -> bool:
    t = normalize_text(text)
    bad = ["out of stock", "нет в наличии", "распродан", "sold out"]
    return not any(x in t for x in bad)


def score_match(query: str, title: str) -> float:
    q_tokens = [t for t in re.split(r"[^a-zа-я0-9]+", normalize_text(query)) if len(t) > 1]
    t = normalize_text(title)
    if not q_tokens:
        return 0.0
    hit = sum(1 for tok in q_tokens if tok in t)
    return hit / len(q_tokens)


def search_ddg(query: str, domain: str, limit: int = 6) -> list[str]:
    q = f"site:{domain} {query}"
    url = f"https://duckduckgo.com/html/?q={quote(q)}"
    r = requests.get(url, headers=UA, timeout=25)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    links = []
    for a in soup.select("a.result__a"):
        href = a.get("href")
        if href and domain in href:
            links.append(href)
        if len(links) >= limit:
            break
    return links


def parse_candidate(url: str) -> tuple[float | None, str, str, bool]:
    r = requests.get(url, headers=UA, timeout=25)
    txt = r.text[:300000]
    title = ""
    m = re.search(r"<title>(.*?)</title>", txt, flags=re.I | re.S)
    if m:
        title = re.sub(r"\s+", " ", m.group(1)).strip()
    currency = detect_currency(txt)
    prices = extract_prices(txt)
    price = min(prices) if prices else None
    return price, currency, title, likely_in_stock(txt)


def best_offer(product: str, domain: str) -> tuple[str, str, str]:
    try:
        urls = search_ddg(product, domain)
    except Exception:
        return "товар не найден", "", ""

    candidates = []
    for u in urls:
        try:
            price, ccy, title, instock = parse_candidate(u)
            if not instock or price is None:
                continue
            rel = score_match(product, title)
            if rel < 0.55:
                continue
            candidates.append((price, ccy, u, rel))
        except Exception:
            continue

    if not candidates:
        return "товар не найден", "", ""

    # Anti-scam: filter too-cheap outliers
    prices = sorted(x[0] for x in candidates)
    median = prices[len(prices) // 2]
    safe = [x for x in candidates if x[0] >= 0.45 * median]
    if safe:
        candidates = safe

    best = min(candidates, key=lambda x: x[0])
    return best[2], f"{best[0]:.2f}", best[1]


def find_columns(header_row: list[str | None]):
    product_col = None
    seller_pairs = []
    for i, h in enumerate(header_row, start=1):
        hs = normalize_text(str(h or ""))
        if "наименование" in hs:
            product_col = i

    for i, h in enumerate(header_row, start=1):
        hs = normalize_text(str(h or ""))
        if "ссылка" in hs and "продавец" in hs:
            m = re.search(r"-\s*([a-zа-я0-9\s]+?)\s*-\s*ссылка", hs, re.I)
            market = (m.group(1).strip() if m else hs).lower()
            price_col = i + 1
            seller_pairs.append((market, i, price_col))

    return product_col, seller_pairs


def market_to_domain(market: str) -> str | None:
    for k, d in MARKET_DOMAIN.items():
        if k in market:
            return d
    return None


def run(input_xlsx: str, output_xlsx: str):
    wb = load_workbook(input_xlsx)
    ws = wb.active

    header_idx = 2
    header = [ws.cell(header_idx, c).value for c in range(1, ws.max_column + 1)]
    product_col, sellers = find_columns(header)
    if not product_col:
        raise RuntimeError("Не найден столбец с названием товара")

    for r in range(header_idx + 1, ws.max_row + 1):
        product = ws.cell(r, product_col).value
        if not product:
            continue
        product = str(product).strip()
        if not product:
            continue

        for market, link_col, price_col in sellers:
            domain = market_to_domain(market)
            if not domain:
                ws.cell(r, link_col).value = "товар не найден"
                ws.cell(r, price_col).value = ""
                continue

            link, price, ccy = best_offer(product, domain)
            ws.cell(r, link_col).value = link
            ws.cell(r, price_col).value = f"{price} {ccy}".strip()

    wb.save(output_xlsx)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    args = p.parse_args()
    run(args.input, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
